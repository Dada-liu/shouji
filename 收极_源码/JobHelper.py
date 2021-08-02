from os import chdir,listdir,rename
import openpyxl
from re import compile,search
from pyperclip import copy


#files_path = "F:\大学\学习\第二学位本科学习\python\自己编写的小插件\标记作业已交人数\作业"
#C:\Users\310\Desktop\JobHelper\江财班级名册.xlsx
#files_path=input(r"请输入作业所在路径(F:\大学\学习\第二学位本科学习\python\自己编写的小插件\标记作业已交人数\作业)：")
#excel_path=input(r"请输入班级 Excel 所在路径(C:\Users\310\Desktop\JobHelper\江财班级名册.xlsx)：")
#sit = eval(input("请输入放在excel中第几列(eg:以1开始的列数)："))


class Op():
    def __init__(self,files_path,excel_path):
        """类的初始化函数"""
        self.files_path = files_path
        self.excel_path = excel_path


    def getstudentlist(self,studentnumberlen):
        """
        获得作业存放路径，返回一个字符串列表,其中元素为 学号 或 卡号  或 姓名,如果返回的是[],则文件夹是空
        studentnumberlen是一个包含两个元素的列表，第一个是学号长度，第二个是卡号长度
        """
        chdir(self.files_path)  # 进入文件储存路径
        filelists = listdir()
        pattern = compile(r"\d+")
        student_numberorname_list = []
        for str_i in filelists:#把文件名中的学号或卡号或姓名读出，形成一个一维列表
            site = pattern.search(str_i)
            if (not site) or (len(str(site.group())) != studentnumberlen[0]
                              and len(str(site.group())) != studentnumberlen[1]):#没有卡号没有学号，就读取姓名或文件名
                site = search(r'[\u4e00-\u9fa5]{2,5}',str_i)
                if len(site.group()) == 5:
                    site = search(r'[\u4e00-\u9fa5]+', str_i)
            student_numberorname_str = site.group()#要对这串数字进行判断，防止是非学号或卡号
            #student_number = eval(student_number_str)
            student_numberorname_list.append(student_numberorname_str)#形成一个元素为 学号 或 卡号  或 姓名 的列表
        return student_numberorname_list#包含了路径中的文件名列表和提取的学号或者姓名表


    def getstudentinfo(self):
        """返回excle中的学生信息"""
        #把获取的路径加文件名分离开
        j = 0
        for i in self.excel_path[-1:0:-1]:
            j += 1
            if i == "\\" or i == "/":
                path1 = self.excel_path[0:len(self.excel_path)-j]#文件路径
                bookname1 = self.excel_path[len(self.excel_path)-j+1:]#excel文件名
                break
        chdir(path1)
        #os.chdir("..\\")#切换到上级目录
        #bookname = "S2授课班 班级名册.xlsx"
        book = openpyxl.load_workbook(bookname1)#用book变量读取文件
        sheet_name = book.sheetnames[0]#.sheetnames 返回excel文件表名的列表
        sheet = book[sheet_name]#通过表名对excel文件的第一个表进行操作
        list1 = []#卡号列表；创建空列表，下面的append()方法才能用
        list2 = []#学号列表
        list3 = []#姓名列表
        for i in range(1, 8):
            if sheet.cell(2, i).value == "卡号":
                for j in range(3, 1000):  # 63后面要换成一个n
                    if sheet.cell(j, i).value == None:
                        break
                    list1.append(str(sheet.cell(j, i).value))
            elif sheet.cell(2, i).value == "学号":
                for j in range(3, 1000):
                    if sheet.cell(j, i).value == None:
                        break
                    list2.append(str(sheet.cell(j, i).value))
            elif sheet.cell(2, i).value == "姓名":
                for j in range(3, 1000):
                    if sheet.cell(j, i).value == None:
                        break
                    list3.append(sheet.cell(j, i).value)
        infolist = list(zip(list3,list2,list1))
        studentnumbenlen = [len(list1[0]),len(list2[0])]
        return [infolist,studentnumbenlen]#这是一个列表，其数据格式为[(姓名1，学号1，卡号1),(姓名2，学号2，卡号2),(),...]

    #把从文件夹中读取的文件名列表和从excel文件中读取的原始列表对比,返回一个0、1的列表，0代表没交或无法识别的文件,student_list是一个以元组为元素的列表
    def compare(self,student_list,infolist):
        length1=len(infolist)
        length2=len(student_list)
        flags = []
        for i in range(len(infolist)):
            flags.append(0)
        i = 0
        for name,studentid,cardid in infolist:#for i in 后只能是可迭代对象
            for j in range(len(student_list)):
                if studentid == student_list[j]or cardid == student_list[j]\
                        or name == student_list[j]:
                    flags[i] = 1
                    break
            i += 1
        return flags

    # 对compare()返回的列表计数，得到值为1的个数
    def counts(self,flags):
        count = 0
        for i in flags:
            if i == 1:
                count = count + 1
        return count

    #获取excel中得到的学生信息，返回一个 未交作业学生名单列表
    def unfinished(self,infolist,flags):
        unfinishedlists = []
        for i in range(len(flags)):
            if flags[i] == 0 :
                unfinishedlists.append(infolist[i][0])
        return unfinishedlists

    #获得从文件中提取的列表，返回无法确定归属的作业文件名
    def unknown(self,student_numberorname_list):
        unknownlists = []
        pattern = compile(r"\d+")
        for i in student_numberorname_list:
            j = pattern.match(i)
            if not j and len(i) > 4:
                unknownlists.append(i)
        return unknownlists

    #复制未交作业的人到剪贴板
    def copy_to_clipboard(self,unfinishedlists):
        agoodprintstrs = unfinishedlists.copy()
        agoodprintstrs.insert(0,"以下是未交作业名单：")
        lists_str = "\n       ".join(agoodprintstrs)
        copy(lists_str)

    #把形成的对比标记列表写入excel文件
    def putflag(self, flags):
        j = 0
        for i in self.excel_path[-1:0:-1]:
            j += 1
            if i == "\\" or i == "/":
                path = self.excel_path[0:len(self.excel_path)-j]#文件路径
                bookname = self.excel_path[len(self.excel_path)-j+1:]#excel文件名
                break
        chdir(path)
        book = openpyxl.load_workbook(bookname)# 用book变量读取文件
        sheet_name = book.sheetnames[0]
        print(path,bookname,sheet_name)
        sheet = book[sheet_name]
        for j in range(1,50):
            print(sheet.cell(2,j).value)

            if sheet.cell(2,j).value == "状态标记":
                for i in range(len(flags)):
                    sheet.cell(i + 3,j).value = flags[i]  # 写入单元格
                break
        try:
            book.save(bookname)
            return 0
        except:
            return 1 #返回 1 代表保存出错


    #获得从文件夹获得的信息student_list和从excel获得的info_list，对文件改名
    def renamefile(self,student_list,info_list):
        chdir(self.files_path)
        oldnamelist = listdir()
        newname_list= []
        #生成一个和文件数量一样大小的列表
        for i in range(len(oldnamelist)):
            newname_list.append(0)
        partten = compile("无法识别文件")
        #得到格式化的新名字列表
        for i in range(len(oldnamelist)):
            if partten.search(oldnamelist[i]) == None:  # 使用search()匹配字符串，如果没有，返回None
                try:
                    newname_list[i] = "无法识别文件+" + oldnamelist[i]
                except FileExistsError:
                    print("错误0")
            else:
                newname_list[i] = oldnamelist[i]
            for name,studentnumber,cardnumber in info_list:

                if student_list[i] == name or student_list[i] == studentnumber or student_list[i] == cardnumber:
                    try:
                        newname_list[i] = str(studentnumber) + "+" + str(name) + "." + oldnamelist[i].split(".")[1]
                    except IndexError:
                        print("错误1")
                    break
        print(oldnamelist,newname_list)
        #给重复文件改名
        chdir(self.files_path)
        for oldname, newname in zip(oldnamelist, newname_list):
            try:
                rename(oldname, newname)
            except FileExistsError:#如果一个文件夹中有多个归属于同一个人的文件，对其进行区分
                x = 0
                savedfilelist = listdir()
                print(savedfilelist)
                str1 = newname.split(".")[0].split("+")[0]
                str2 = newname.split(".")[0].split("+")[1]
                for i in savedfilelist:
                        if search(str1,i) != None or search(str2,i) != None:
                            if search("重复文件" + str(x), i) != None:
                                x +=1
                            """
                            for x in range(100):
                                if search("重复文件"+ str(x),i) != None:
                                    print(search("重复文件"+ str(x),i))
                                    print(x)
                                else:
                                    break

                            """
                try:
                    rename(oldname,newname.split(".")[0] + "+" + "重复文件"+ str(x) + "." + newname.split(".")[1])
                except:
                    print("错误3")

def renamef(file_path,excel_path):
    object1 = Op(file_path,excel_path)
    info_list = object1.getstudentinfo()
    studentlist = object1.getstudentlist(info_list[1])
    object1.renamefile(studentlist,info_list[0])

def greatoutcome(file_path,excel_path):
    object2 = Op(file_path,excel_path)
    info_list = object2.getstudentinfo()
    student_list = object2.getstudentlist(info_list[1])

    if student_list == []:
        aerro1 = 1
    else:
        aerro1 = 0
    flags = object2.compare(student_list, info_list[0])
    aerro = object2.putflag(flags)
    #finished_sum = s2.counts(flags)
    unfinishedlist = object2.unfinished(info_list[0], flags)
    unknownlist = object2.unknown(student_list)
    #object2.copy_to_clipboard(unfinishedlist)
    return [unfinishedlist,unknownlist,aerro,aerro1]

def copytoclipboard1(file_path,excel_path):
    object3 = Op(file_path,excel_path)
    info_list = object3.getstudentinfo()
    student_list = object3.getstudentlist(info_list[1])
    flags = object3.compare(student_list, info_list[0])
    unfinishedlist = object3.unfinished(info_list[0], flags)
    print(unfinishedlist)
    object3.copy_to_clipboard(unfinishedlist)

def copytoclipboard2(file_path,excel_path):
    object3 = Op(file_path,excel_path)
    info_list = object3.getstudentinfo()
    student_list = object3.getstudentlist(info_list[1])
    unknownlist = object3.unknown(student_list)
    print(unknownlist)
    object3.copy_to_clipboard(unknownlist)

"""
s2 = Op(files_path,excel_path)
student_list = s2.getstudentlist()
infolist = s2.getstudentinfo()
flags = s2.compare(student_list,infolist)
finished_sum = s2.counts(flags)
unfinishedlist = s2.unfinished(infolist,flags)
unknownlist = s2.unknown(student_list)
s2.copy_to_clipboard(unfinishedlist)
"""